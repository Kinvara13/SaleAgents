from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, status, BackgroundTasks
from sqlalchemy.orm import Session
from pathlib import Path
from uuid import uuid4
import zipfile, tempfile, os, shutil

from app.db.session import get_db
from app.schemas.parsing import ParsingSectionSummary, ParsingSectionDetail, ParsingSectionUpdateRequest
from app.services.llm_parsing_client import llm_parsing_client
from app.models.parsing_section import ParsingSection
from app.models.project import Project
from app.core.config import settings

router = APIRouter()

ALLOWED_EXTENSIONS = {".pdf", ".doc", ".docx", ".txt", ".ppt", ".pptx", ".xls", ".xlsx", ".zip"}
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB
UNZIP_DIR = Path(settings.storage_path or Path(__file__).resolve().parents[4] / "storage") / "unzip_temp"


def _extract_text_from_file(file_path: Path, ext: str) -> str:
    """Extract text content from a file."""
    try:
        if ext == ".txt":
            return file_path.read_text(encoding="utf-8", errors="replace")
        elif ext == ".docx":
            try:
                from docx import Document
                doc = Document(str(file_path))
                return "\n".join([p.text for p in doc.paragraphs])
            except Exception:
                return ""
        elif ext == ".xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(file_path), data_only=True)
                lines = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        line = " ".join([str(c) if c else "" for c in row])
                        if line.strip():
                            lines.append(line)
                return "\n".join(lines)
            except Exception:
                return ""
        elif ext == ".pdf":
            try:
                import pypdf
                reader = pypdf.PdfReader(str(file_path))
                text = ""
                for page in reader.pages:
                    text += page.extract_text() or ""
                return text
            except Exception:
                try:
                    import pdfplumber
                    text = ""
                    with pdfplumber.open(str(file_path)) as pdf:
                        for page in pdf.pages:
                            text += page.extract_text() or ""
                    return text
                except Exception:
                    return ""
        else:
            # Fallback: read as text
            return file_path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        return f"[文件读取失败: {e}]"


def _parse_uploaded_file(project_id: str, file_path: Path, filename: str, db: Session) -> list[dict]:
    """Parse a single uploaded file and return section dicts."""
    ext = Path(filename).suffix.lower()
    text = _extract_text_from_file(file_path, ext)
    
    if not text or len(text.strip()) < 50:
        return []
    
    # Use LLM to extract tender info
    tender_info = llm_parsing_client.extract_tender_fields(text)
    
    sections = []
    
    # Create sections based on what was parsed
    if tender_info:
        # Parse scoring rules (评分重点) into individual sections
        scoring_text = tender_info.get("评分重点", {}).get("value", "") or tender_info.get("评分重点", "")
        if scoring_text and len(scoring_text) > 20:
            sections.append({
                "project_id": project_id,
                "section_name": "评分规则解析",
                "section_type": "评审",
                "content": scoring_text,
                "is_star_item": True,
                "source_file": filename,
            })
        
        # Technical requirements
        tech_req = tender_info.get("技术要求", {}).get("value", "") or tender_info.get("技术要求", "")
        if tech_req and len(tech_req) > 20:
            sections.append({
                "project_id": project_id,
                "section_name": "技术要求",
                "section_type": "评审",
                "content": tech_req,
                "is_star_item": True,
                "source_file": filename,
            })
        
        # Business sections
        for name in ["商务偏离表", "应答承诺函", "授权委托书", "营业执照"]:
            sections.append({
                "project_id": project_id,
                "section_name": name,
                "section_type": "商务",
                "content": f"【{name}】\n\n解析自招标文件 {filename}，请根据招标要求及公司实际情况填写。\n\n参考信息：{tender_info.get('必备资质', {}).get('value', '待补充')}",
                "is_star_item": name in ["商务偏离表", "应答承诺函"],
                "source_file": filename,
            })
        
        # Technical sections
        for name in ["技术条款偏离表", "CMMI证书", "计算机软件著作权证书", "项目案例", "自查确认单"]:
            sections.append({
                "project_id": project_id,
                "section_name": name,
                "section_type": "技术",
                "content": f"【{name}】\n\n解析自招标文件 {filename}，请根据技术规范书要求填写。",
                "is_star_item": name in ["技术条款偏离表", "项目案例"],
                "source_file": filename,
            })
    else:
        # Fallback: create generic sections without LLM
        for name in ["商务偏离表", "应答承诺函", "授权委托书", "营业执照", "技术条款偏离表"]:
            sections.append({
                "project_id": project_id,
                "section_name": name,
                "section_type": "商务" if name in ["商务偏离表", "应答承诺函", "授权委托书", "营业执照"] else "技术",
                "content": f"【{name}】\n\n上传文件：{filename}\n\n（请在下方编辑器中填写内容）",
                "is_star_item": False,
                "source_file": filename,
            })
    
    return sections


def _handle_zip(project_id: str, zip_path: Path, db: Session) -> list[dict]:
    """Extract and parse all files inside a ZIP archive."""
    import tempfile
    all_sections = []
    
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(UNZIP_DIR / project_id)
        
        # Walk extracted files
        for root, dirs, files in os.walk(UNZIP_DIR / project_id):
            for fname in files:
                fpath = Path(root) / fname
                ext = fpath.suffix.lower()
                if ext in ALLOWED_EXTENSIONS and ext != ".zip":
                    sections = _parse_uploaded_file(project_id, fpath, fname, db)
                    all_sections.extend(sections)
    except Exception as e:
        pass
    
    # Cleanup
    shutil.rmtree(UNZIP_DIR / project_id, ignore_errors=True)
    return all_sections


@router.post("/{project_id}/upload")
async def upload_and_parse(
    project_id: str,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> dict:
    """上传招标文件（支持ZIP包）并触发后台异步解析"""
    filename = file.filename or "unknown"
    ext = Path(filename).suffix.lower()

    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"不支持的文件类型，仅支持: {', '.join(ALLOWED_EXTENSIONS)}",
        )

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="文件大小超过100MB限制")

    # Ensure project exists
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")

    upload_dir = Path(settings.storage_path or Path(__file__).resolve().parents[4] / "storage") / "projects" / project_id / "bid_documents"
    upload_dir.mkdir(parents=True, exist_ok=True)

    if ext == ".zip":
        # Handle ZIP archive
        zip_path = upload_dir / filename
        with open(zip_path, "wb") as f:
            f.write(content)
        background_tasks.add_task(_handle_zip_async, project_id, zip_path)
    else:
        # Handle single file
        file_path = upload_dir / filename
        with open(file_path, "wb") as f:
            f.write(content)
        background_tasks.add_task(_parse_single_file_async, project_id, file_path, filename)

    return {"status": "processing", "message": "文件已接收，后台解析任务已启动"}

def _handle_zip_async(project_id: str, zip_path: Path):
    from app.db.session import SessionLocal
    db = SessionLocal()
    try:
        project = db.query(Project).filter(Project.id == project_id).first()
        if project:
            project.status = "解析中"
            project.parse_status = "解析中"
            if isinstance(project.node_status, dict):
                project.node_status["parsing"] = "processing"
            else:
                project.node_status = {"parsing": "processing"}
            db.commit()
            
        sections = _handle_zip(project_id, zip_path, db)
        _save_parsed_sections(project_id, sections, db)
        zip_path.unlink(missing_ok=True)
        
        if project:
            project.status = "解析完成"
            project.parse_status = "已解析"
            if isinstance(project.node_status, dict):
                project.node_status["parsing"] = "completed"
            db.commit()
    except Exception as e:
        import logging
        logging.error(f"Zip parsing failed for project {project_id}: {e}")
        db.rollback()
        project = db.query(Project).filter(Project.id == project_id).first()
        if project:
            project.status = "解析失败"
            project.parse_status = "解析失败"
            if isinstance(project.node_status, dict):
                project.node_status["parsing"] = "failed"
            db.commit()
    finally:
        db.close()

def _parse_single_file_async(project_id: str, file_path: Path, filename: str):
    from app.db.session import SessionLocal
    db = SessionLocal()
    try:
        project = db.query(Project).filter(Project.id == project_id).first()
        if project:
            project.status = "解析中"
            project.parse_status = "解析中"
            if isinstance(project.node_status, dict):
                project.node_status["parsing"] = "processing"
            else:
                project.node_status = {"parsing": "processing"}
            db.commit()
            
        sections = _parse_uploaded_file(project_id, file_path, filename, db)
        _save_parsed_sections(project_id, sections, db)
        
        if project:
            project.status = "解析完成"
            project.parse_status = "已解析"
            if isinstance(project.node_status, dict):
                project.node_status["parsing"] = "completed"
            db.commit()
    except Exception as e:
        import logging
        logging.error(f"Single file parsing failed for project {project_id}: {e}")
        db.rollback()
        project = db.query(Project).filter(Project.id == project_id).first()
        if project:
            project.status = "解析失败"
            project.parse_status = "解析失败"
            if isinstance(project.node_status, dict):
                project.node_status["parsing"] = "failed"
            db.commit()
    finally:
        db.close()

def _save_parsed_sections(project_id: str, sections: list[dict], db: Session):
    db.query(ParsingSection).filter(ParsingSection.project_id == project_id).delete()

    for sec in sections:
        section = ParsingSection(
            id=f"sec_{uuid4().hex[:12]}",
            project_id=sec["project_id"],
            section_name=sec["section_name"],
            section_type=sec["section_type"],
            content=sec["content"],
            is_star_item=sec["is_star_item"],
            source_file=sec["source_file"],
        )
        db.add(section)

    db.commit()


@router.get("/{project_id}/sections", response_model=list[ParsingSectionSummary])
def get_sections(project_id: str, db: Session = Depends(get_db)) -> list[ParsingSectionSummary]:
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    sections = (
        db.query(ParsingSection)
        .filter(ParsingSection.project_id == project_id)
        .order_by(ParsingSection.section_type, ParsingSection.id)
        .all()
    )
    return [ParsingSectionSummary.model_validate(s) for s in sections]


@router.get("/{project_id}/sections/{section_id}", response_model=ParsingSectionDetail)
def get_section_detail(
    project_id: str,
    section_id: str,
    db: Session = Depends(get_db),
) -> ParsingSectionDetail:
    section = (
        db.query(ParsingSection)
        .filter(ParsingSection.id == section_id, ParsingSection.project_id == project_id)
        .first()
    )
    if not section:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="章节不存在")
    return ParsingSectionDetail.model_validate(section)


@router.patch("/{project_id}/sections/{section_id}", response_model=ParsingSectionDetail)
def patch_section(
    project_id: str,
    section_id: str,
    payload: ParsingSectionUpdateRequest,
    db: Session = Depends(get_db),
) -> ParsingSectionDetail:
    section = (
        db.query(ParsingSection)
        .filter(ParsingSection.id == section_id, ParsingSection.project_id == project_id)
        .first()
    )
    if not section:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="章节不存在")
    if payload.content is not None:
        section.content = payload.content
    db.commit()
    db.refresh(section)
    return ParsingSectionDetail.model_validate(section)